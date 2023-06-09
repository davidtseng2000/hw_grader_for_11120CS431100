import shutil
import os
from os import walk
from os.path import join
import re
from colorama import init, Fore
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

######################
# 一些可能需要修改的變數
######################
global MyPath # 資料夾路徑
global num_begin # 學號開頭 (如: 106, 107, 108, 109, 110, 111)
global excel_file # excel 檔案名稱 (如: 計算方法設計 成績.xlsx)
global excel_sheet # excel 檔案中的哪個分頁 (如: "計算方法設計 成績.xlsx" 中的 "作業5")



##############################################################
# Step1: 一開始先列出所有繳交作業的名單，並且在 excel 表格中做記號
##############################################################

# (1-1) 列出所有繳交作業的名單
MyPath  = './'
source = walk(MyPath)
folder, subfolders, filenames = next(source)
num_begin = ['106', '107', '108', '109', '110', '111', 'x']
all_submit = set([str(number.split(' ')[0]) for number in subfolders if any(begin in number for begin in num_begin)])
print(Fore.WHITE + f"所有繳交名單:\n{all_submit}")
print(f"共 {len(all_submit)} 人\n")

# (1-2) 在 excel 表格中做記號
try:
  excel_file = "計算方法設計 成績.xlsx"
  wb = openpyxl.load_workbook(excel_file)
  excel_sheet = "作業5"
  target_excel_sheet = wb[excel_sheet]
  # 學號是第一行
  student_id_column_index = 1
  # 有交作業者用淺藍色標示
  fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
  for column in target_excel_sheet.iter_cols(min_row=2, max_row=target_excel_sheet.max_row, min_col=student_id_column_index, max_col=student_id_column_index):
      for cell in column:
          student_id = str(cell.value)
          if student_id in all_submit:
              cell.fill = fill
  # 保存修改后的 Excel 文件
  wb.save("計算方法設計 成績_updated.xlsx")

except:
  print("未讀取到 excel 檔案")

##################################
# Step2: 列出檔名/格式可能有誤的名單
##################################

MyPath  = './'
# Keyword 紀錄檔名規定 (如: hw4_bnb, hw5_pns)
Keyword = input(Fore.BLUE + "請輸入規定之檔名: (如: hw4_bnb, hw5_pns)\n")

# 預設學生繳交檔案格式為 .cpp 或 .py
# no_cpp_py 紀錄沒交 .cpp 或 .py 檔案的學生
# no_keyword 紀錄檔名非規定格式的學生 (如: hw4_bnb, hw5_pns)
no_cpp_py = []
no_keyword = []
for folder, subfolders, filenames in walk(MyPath):
  if any(begin in folder for begin in num_begin):
    if (not any(file.endswith(".cpp") for file in filenames)) and (not any(file.endswith(".c") for file in filenames)) and (not any(file.endswith(".py") for file in filenames)) and ("\\" not in folder):
      no_cpp_py.append(folder)   
    if (not any(Keyword in file for file in filenames)) and ("\\" not in folder):
      no_keyword.append(folder)

no_cpp_py = set(no_cpp_py).difference({"./"})
no_keyword = set(no_keyword).difference({"./"})

equal_len = len(str(no_cpp_py.union(no_keyword)))
print(Fore.WHITE + "=" * equal_len)
print("以下同學沒有交 .c, .cpp 或 .py 檔:")
print(no_cpp_py)
print("\n")
print(f"以下同學沒有交符合 {Keyword} 命名格式的檔案:")
print(no_keyword)
print("\n")
print("(統整)所有繳交格式不符合:")
print(no_cpp_py.union(no_keyword))
print("=" * equal_len)


##################################
# Step3: 整理資料夾 (.cpp/.py 分類)
##################################
exten_count = 0
# 建立一個蒐集所有副檔名為 FileExten(如: .cpp/.py) 的檔案，集中在一個資料夾
# FileExten = input(Fore.BLUE + '請輸入副檔名關鍵字: (如: .cpp/.py)\n')
def organize_folder(FileExten):
  if not os.path.exists(FileExten):
    os.mkdir(FileExten)

  MyPath = './'
  move = 0
  for folder, subfolders, filenames in walk(MyPath):
    id = re.split("./| ", folder)[1]
    for i in filenames:
      
      FullPath = join(folder, i) # 獲取檔案完整路徑
      FileName = join(i) # 獲取檔案名稱 

      try:
        # 改名      
        if (Keyword in FileName) and FileName.endswith(FileExten):
          shutil.copy(FullPath, MyPath + '/' + FileExten)
          NewFileName = Keyword + "_" + id + FileExten
          os.rename(MyPath + '/' + FileExten + '/' + FileName, MyPath + '/' + FileExten + '/' + NewFileName)            
          print (Fore.WHITE + '成功將', FileName, '改名為', NewFileName, '並移動至', FileExten, '資料夾')
          move = move + 1   
      except Exception:
        pass

  print("總共移動了 " + str(move) + " 個 " + FileExten + " 類型的檔案" + '\n')

  # 紀錄目前 .cpp/.py 資料夾中的檔案 (學號)
  MyPath  = f'./{FileExten}'
  global exten_count
  globals()['exten_list'+str(exten_count)] = set([file.lstrip(Keyword).rstrip(FileExten) for folder, subfolders, filenames in walk(MyPath) for file in filenames])
  # print(f"{globals()['exten_list'+str(exten_count)]}\n")
  exten_count = exten_count + 1
  
# 預設繳交的程式為 C++, C 或 Python
print('')
organize_folder('.cpp')
organize_folder('.c')
organize_folder('.py')